#!/usr/bin/python
# -*- Coding: utf-8 -*-
#%%
import numpy as np
import pandas as pd
import os
import glob
import re
import openpyxl
from openpyxl import Workbook, load_workbook

# ファイルの位置の一つ上のフォルダのパスを取得
dirpath = os.path.dirname(os.path.abspath(__file__))

# 結果フォルダのエクセルファイル名を取得(絶対パス)
result_paths = glob.glob(dirpath+"/2.結果/*.xlsx")


# ====== このエリアはいらないといえば、いらないが、ファイルを番号順に読み込む保証は全くないため、きちんと番号順に並び替える ======
# ファイル名だけにする
result_names = []
for f in result_paths:
    result_names.append(os.path.split(f)[1])

# ファイル名から10進数の数字だけ抜き出す
# re.sub(正規表現パターン, 置換後文字列, 置換したい文字列)
# \D : 10進数でない任意の文字。（全角数字等を含む）
result_numbers = []
for name in result_names: 
    num = re.sub("\\D", "", name)
    result_numbers.append(num)


# 抜き出した数字を元にファイル名を順番に並び替えする

# まずは、絶対パスのファイル名, ファイル名, 番号のデータフレームを作成する。
# リストは行ベクトルになっていて、それをくっつけると、1行目にpaths, 2行目にnames,...とな流ので、
# 1列目にpaths,...となるように転置する。(.T)
result_name_df = pd.DataFrame([result_paths, result_names, result_numbers]).T
result_name_df.columns = ['path', 'name', 'number']

# このデータフレームをnumberをキーにして並び替える。
result_name_df.sort_values('number',ascending=True)
result_name_df = result_name_df.reset_index(drop=True)

# ソートされたファイル名だけのものもリストとして取り出しておく
result_paths_sorted = list(result_name_df['path'])
# =============================================================

#%%
# ----------------- 本題 ---------------------
# ファイルを順番に読み込んでいく。
    
wbr_list = [load_workbook(path) for path in result_paths_sorted]

# 全く同じコードの別表現(上はリストの内包表記)
#wbr_list = []
#for path in result_paths_sorted:
#    wbr_list.append(load_workbook(path))

# シートを選択
sheet_name = "集計シート（記入不要）"
sheets = [wbr[sheet_name] for wbr in wbr_list]

# シートの中のコピーする行(D=3)を選択し、内容を取得
copy_cells = []
for sheet in sheets:
    tmp = []
    for i, cell_obj in enumerate(list(sheet.columns)[3]):
        # 最初の2つのセルは不要なので落としておく
        if i==0 or i==1:
            continue
        tmp.append(cell_obj.value)
    copy_cells.append(tmp)


#%%
# コピー先のファイルを開く
wbs = load_workbook(dirpath+'/3.フィードバック/結果ファイル.xlsx')

sheet_name_s = "全店結果一覧表"
sheet = wbs[sheet_name_s]

for i, datas in enumerate(copy_cells):
    for j, data in enumerate(datas):
        sheet.cell(column=3+i, row=4+j, value=data)

wbs.save(dirpath+'/3.フィードバック/結果ファイル_.xlsx')

